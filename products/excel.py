from decimal import Decimal, InvalidOperation

from openpyxl import Workbook, load_workbook

from masters.models import PriceLevel, Tax
from .models import Category, Brand, Unit, Product, ProductPrice

EXPORT_HEADERS = [
    'name', 'barcode', 'sku', 'category', 'brand', 'unit', 'tax',
    'purchase_price', 'retail_price', 'minimum_stock',
    'expiry_date', 'manufacture_date', 'status',
]


def export_products_workbook(products):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Products'
    ws.append(EXPORT_HEADERS)
    for p in products:
        ws.append([
            p.name, p.barcode or '', p.sku or '',
            p.category.name if p.category else '',
            p.brand.name if p.brand else '',
            p.unit.name if p.unit else '',
            p.tax.name if p.tax else '',
            p.purchase_price, p.retail_price, p.minimum_stock,
            p.expiry_date.isoformat() if p.expiry_date else '',
            p.manufacture_date.isoformat() if p.manufacture_date else '',
            p.status,
        ])
    return wb


def _clean(value):
    if value is None:
        return ''
    return str(value).strip()


def import_products_from_workbook(wb):
    """Row-by-row upsert, matched by barcode first then sku. Returns
    (created_count, updated_count, errors) where errors is a list of
    (row_number, message) — a bad row is skipped, not fatal to the batch."""
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return 0, 0, []

    header = [_clean(h).lower() for h in rows[0]]
    col = {name: idx for idx, name in enumerate(header)}
    if 'name' not in col:
        return 0, 0, [(1, "Header row must include a 'name' column.")]

    created = 0
    updated = 0
    errors = []

    def get(row, key, default=''):
        idx = col.get(key)
        if idx is None or idx >= len(row):
            return default
        value = row[idx]
        return default if value is None else value

    for row_num, row in enumerate(rows[1:], start=2):
        if row is None or not any(row):
            continue
        name = _clean(get(row, 'name'))
        if not name:
            errors.append((row_num, "Missing product name — row skipped."))
            continue

        barcode = _clean(get(row, 'barcode')) or None
        sku = _clean(get(row, 'sku')) or None

        product = None
        if barcode:
            product = Product.objects.filter(barcode=barcode).first()
        if product is None and sku:
            product = Product.objects.filter(sku=sku).first()

        is_new = product is None
        if is_new:
            product = Product(name=name, barcode=barcode, sku=sku)
        else:
            product.name = name

        category_name = _clean(get(row, 'category'))
        if category_name:
            product.category, _ = Category.objects.get_or_create(name=category_name)

        brand_name = _clean(get(row, 'brand'))
        if brand_name:
            product.brand, _ = Brand.objects.get_or_create(name=brand_name)

        unit_name = _clean(get(row, 'unit'))
        if unit_name:
            product.unit, _ = Unit.objects.get_or_create(name=unit_name)

        tax_name = _clean(get(row, 'tax'))
        if tax_name:
            tax = Tax.objects.filter(name=tax_name).first()
            if tax:
                product.tax = tax
            else:
                errors.append((row_num, f"Tax '{tax_name}' not found — left unchanged."))

        try:
            purchase_price = get(row, 'purchase_price')
            if purchase_price not in ('', None):
                product.purchase_price = Decimal(str(purchase_price))
            minimum_stock = get(row, 'minimum_stock')
            if minimum_stock not in ('', None):
                product.minimum_stock = Decimal(str(minimum_stock))
        except InvalidOperation:
            errors.append((row_num, "Invalid number in purchase_price/minimum_stock — row skipped."))
            continue

        expiry = _clean(get(row, 'expiry_date'))
        if expiry:
            product.expiry_date = expiry
        manufacture = _clean(get(row, 'manufacture_date'))
        if manufacture:
            product.manufacture_date = manufacture

        status = _clean(get(row, 'status')).lower()
        if status in dict(Product.Status.choices):
            product.status = status

        try:
            product.full_clean(exclude=['stock'])
            product.save()
        except Exception as exc:
            errors.append((row_num, str(exc)))
            continue

        retail_price = get(row, 'retail_price')
        if retail_price not in ('', None):
            try:
                level, _ = PriceLevel.objects.get_or_create(name='Retail', defaults={'is_default': True})
                ProductPrice.objects.update_or_create(
                    product=product, price_level=level, defaults={'price': Decimal(str(retail_price))}
                )
            except InvalidOperation:
                errors.append((row_num, "Invalid retail_price — product saved without updating price."))

        if is_new:
            created += 1
        else:
            updated += 1

    return created, updated, errors
