from decimal import Decimal
from io import BytesIO

from django.test import TestCase
from openpyxl import load_workbook

from .models import Category, Brand, Unit, Product
from .excel import export_products_workbook, import_products_from_workbook
from .barcodes import generate_barcode_data_uri


class BarcodeGenerationTests(TestCase):
    def test_generates_a_png_data_uri(self):
        uri = generate_barcode_data_uri('PRD00000001')
        self.assertTrue(uri.startswith('data:image/png;base64,'))
        self.assertGreater(len(uri), 100)


class ProductExcelRoundTripTests(TestCase):
    """Export must produce a sheet import_products_from_workbook can read
    straight back in — the whole point of round-trippable Excel data."""

    def test_export_then_import_updates_existing_product(self):
        category = Category.objects.create(name='Excel Test Category')
        brand = Brand.objects.create(name='Excel Test Brand')
        unit = Unit.objects.create(name='Excel Test Unit', short_name='u')
        product = Product.objects.create(
            name='Excel Test Product', barcode='EXCEL-001', category=category, brand=brand, unit=unit,
            purchase_price=Decimal('20'), minimum_stock=Decimal('3'),
        )

        wb = export_products_workbook(Product.objects.filter(pk=product.pk))
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        reloaded_wb = load_workbook(filename=buffer)
        ws = reloaded_wb.active
        header = [c.value for c in ws[1]]
        name_idx = header.index('name')
        ws.cell(row=2, column=name_idx + 1, value='Excel Test Product (Renamed)')

        created, updated, errors = import_products_from_workbook(reloaded_wb)

        self.assertEqual(errors, [])
        self.assertEqual((created, updated), (0, 1), 'barcode match should update, not create a duplicate')
        product.refresh_from_db()
        self.assertEqual(product.name, 'Excel Test Product (Renamed)')

    def test_import_creates_new_product_with_new_barcode(self):
        wb = export_products_workbook(Product.objects.none())
        ws = wb.active
        ws.append([
            'Brand New Product', 'EXCEL-NEW-001', '', 'Excel Import Category', 'Excel Import Brand', 'Excel Import Unit',
            '', '15', '25', '2', '', '', 'active',
        ])

        before_count = Product.objects.count()
        created, updated, errors = import_products_from_workbook(wb)

        self.assertEqual(errors, [])
        self.assertEqual((created, updated), (1, 0))
        self.assertEqual(Product.objects.count(), before_count + 1)
        new_product = Product.objects.get(barcode='EXCEL-NEW-001')
        self.assertEqual(new_product.retail_price, Decimal('25'))

    def test_missing_name_column_is_reported_not_fatal(self):
        wb = export_products_workbook(Product.objects.none())
        ws = wb.active
        ws.append(['', 'EXCEL-BAD-001'])

        created, updated, errors = import_products_from_workbook(wb)
        self.assertEqual((created, updated), (0, 0))
        self.assertEqual(len(errors), 1)
